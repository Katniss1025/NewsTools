import imaplib
import email
from email.header import decode_header
from email.utils import parseaddr, parsedate_to_datetime
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import io
import re 
from datetime import timedelta
import pandas as pd
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
import os

# 连接并登陆邮箱服务器
def connect_mail(server, port, address, password):
    # 连接邮箱
    try:
        mail = imaplib.IMAP4_SSL(server, port)
        print("Connection successful!")
    except Exception as e:
        print(f"Connection failed: {e}")
    # 登陆邮箱
    try:
        mail.login(address, password)
        print("Email Login successful!")
    except Exception as e:
        print(f"Email login failed: {e}")
    return mail, mail.state

def select_mail(mail):
    status, msgs = mail.select('INBOX')
    return status, msgs

def search_mail(mail, from_address):
    # 搜索未读邮件并指定UTF-8编码
    status, messages = mail.search(None, 'UNSEEN', 'CHARSET', 'UTF-8', 'FROM', f'"{from_address}"')
    return status, messages


def decode_subject(encoded_subject):
    """解码邮件标题"""
    decoded_parts = decode_header(encoded_subject)
    subject_parts = []
    for part, encoding in decoded_parts:
        if isinstance(part, bytes):
            subject_parts.append(part.decode(encoding or 'utf-8', errors='replace'))
        else:
            subject_parts.append(part)
    return ''.join(subject_parts)


def process_email(mail, email_id):
    """处理单封邮件，提取标题和表格附件"""
    # 获取邮件完整内容
    status, data = mail.fetch(email_id, '(RFC822)')
    if status != 'OK':
        return None, "获取邮件内容失败"

    # 解析邮件
    msg = email.message_from_bytes(data[0][1])
    subject = decode_subject(msg['Subject'])
    from_email = parseaddr(msg['From'])[1]
    
    # 获取邮件接收时间
    received_date = parsedate_to_datetime(msg.get('Date'))
    # 计算前一天日期
    previous_day = (received_date - timedelta(days=1))
    time_slot = previous_day.strftime('%Y-%m-%d') + '-' + received_date.strftime('%Y-%m-%d')

    tables = []

    # 遍历邮件部分，查找表格附件
    for part in msg.walk():
        if part.get_content_maintype() == 'multipart':
            continue
        if part.get('Content-Disposition') is None:
            continue

        # 获取附件文件名
        filename = part.get_filename()
        if not filename:
            continue

        # 解码文件名
        # 修复category判断逻辑
        decoded_filename = decode_subject(filename)
        # 问题在这里：当前的条件判断写法是错误的，'境外'或'国外'被当作了单独的布尔表达式
        # 正确的写法应该是使用 'in' 操作符来检查字符串是否包含特定子串
        if '境外' in decoded_filename or '国外' in decoded_filename:
            category = 1
        elif '境内' in decoded_filename or '国内' in decoded_filename:
            category = 0
        else:
            # 可以设置一个默认值或根据需求处理
            category = None  # 或者根据实际情况设置为0或1
            print('无法识别分类')
            exit()

        # 检查是否为表格文件
        if any(decoded_filename.endswith(ext) for ext in ['.xlsx', '.xls', '.csv']):
            # 读取附件内容
            attachment_content = part.get_payload(decode=True)
            if attachment_content:
                df = parse_table_attachment(attachment_content, decoded_filename)
                if df is not None:
                    tables.append({
                        'filename': decoded_filename,
                        'dataframe': df,
                        'content': df.to_string()  # 转换为字符串以便打印
                    })
                else: exit()

    return {
        'email_id': email_id,
        'subject': subject,
        'from_email': from_email,
        'received_date': received_date,
        'time_slot': time_slot, 
        'tables': tables,
        'category': category
    }, "【邮件解析成功】"


def parse_table_attachment(attachment_content, filename):
    """解析表格附件内容"""
    try:
        # 支持Excel和CSV格式
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            # 使用openpyxl引擎读取Excel文件以获取超链接
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(attachment_content), data_only=False)  # data_only=False才能读取公式
            
            # 只处理名为"去重后文章"的sheet
            if "去重后文章" in wb.sheetnames:
                ws = wb["去重后文章"]
                print(f"【选择工作簿: 去重后文章】")
            else:
                print(f"【未找到名为'去重后文章'工作簿】")
                return None
            
            # 获取表头行（第二行，索引1）
            headers = [cell.value for cell in ws[2]]
            
            # 修改：查找'文章链接地址'列索引而不是'原文链接'
            link_col_index = None
            for i, header in enumerate(headers):
                if header == '文章链接地址':
                    link_col_index = i
                    break
            
            # 读取数据行（从第三行开始）
            data = []
            for row in ws.iter_rows(min_row=3, values_only=False):
                row_data = [cell.value for cell in row]
                # 如果找到链接列且单元格包含超链接
                if link_col_index is not None and row[link_col_index].hyperlink:
                    row_data[link_col_index] = row[link_col_index].hyperlink.target
                data.append(row_data)
            
            # 创建DataFrame
            df = pd.DataFrame(data, columns=headers)
        elif filename.endswith('.csv'):
            # 读取CSV文件
            df = pd.read_csv(io.BytesIO(attachment_content), header=1)
            
            # 修改：查找'文章链接地址'列而不是'原文链接'
            if '文章链接地址' in df.columns:
                # 解析CSV中的超链接格式（如=HYPERLINK("url","text")）
                url_pattern = re.compile(r'HYPERLINK\("([^"]+)",')
                df['文章链接地址'] = df['文章链接地址'].apply(lambda x: url_pattern.search(str(x)).group(1) if url_pattern.search(str(x)) else x)
        else:
            return None, f"不支持的表格格式: {filename}"
        print(f"【成功解析表格: {filename}, 共{len(df)}行数据】")
        return df
    except Exception as e:
        print(f"【解析表格失败: {str(e)}】")
        return None


# 添加发送附件邮件的函数
def send_email_with_attachment(smtp_server, smtp_port, from_addr, password, to_addr, subject, body, attachment_path):
    """
    发送带有附件的邮件
    """
    try:
        # 创建邮件对象
        msg = MIMEMultipart()
        msg['From'] = from_addr
        msg['To'] = to_addr
        msg['Subject'] = subject
        
        # 添加邮件正文
        msg.attach(MIMEText(body, 'plain', 'utf-8'))
        
        # 添加附件
        with open(attachment_path, 'rb') as file:
            part = MIMEBase('application', 'vnd.ms-excel')
            part.set_payload(file.read())
            encoders.encode_base64(part)
            filename = os.path.basename(attachment_path)
            part.add_header('Content-Disposition', 'attachment', 
                filename=("utf-8", "", filename))            
            msg.attach(part)
        
        # 连接SMTP服务器并发送邮件
        with smtplib.SMTP_SSL(smtp_server, smtp_port) as server:
            server.ehlo()
            server.login(from_addr, password)
            print("【SMTP登录成功】")
            server.send_message(msg)
            print(f"【邮件已发送至{to_addr}】")
            server.quit()
        
        return True
    except Exception as e:
        print(f"【邮件发送失败: {str(e)}】")
        return False


def mark_email_as_read(mail, email_id):
    """
    将指定邮件标记为已读
    
    参数:
        mail: imaplib.IMAP4_SSL 对象
        email_id: 邮件ID字符串
    
    返回:
        bool: 如果标记成功返回True，否则返回False
    """
    try:
        # 使用 imaplib 的 STORE 命令和 +FLAGS 来添加已读标记
        mail.store(email_id, '+FLAGS', '\Seen')
        print(f"【邮件 {email_id} 已标记为已读】")
        return True
    except Exception as e:
        print(f"【标记邮件 {email_id} 为已读失败: {str(e)}】")
        return False